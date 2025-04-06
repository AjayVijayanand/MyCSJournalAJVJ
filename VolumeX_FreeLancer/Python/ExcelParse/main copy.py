import openpyxl as xl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl_image_loader import SheetImageLoader

logging.basicConfig(filename="newfile.log",
                    format='%(asctime)s %(message)s',
                    filemode='w')

# Creating an object
logger = logging.getLogger()

# Setting the threshold of logger to DEBUG
logger.setLevel(logging.DEBUG)
wb = xl.load_workbook("ch155-001-018-9780323884051.xlsx")
ws = wb.active

print(f'The number of rows are {ws.max_row}')
print(f'The number of columns are {ws.max_column}')

print("Styling the data!")
for COL in range(1, ws.max_column+1):
    for ROW in range(1, ws.max_row+1):
        ws.cell(row = ROW, column = COL).border = Border(
            top = Side(style = 'thin'),
            left = Side(style = 'thin'),
            right = Side(style = 'thin'),
            bottom = Side(style = 'thin')
        )
        if ROW != 1:
            ws.cell(row = ROW, column = COL).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
        else:
            ws.cell(row = ROW, column = COL).alignment = Alignment(horizontal = "center", vertical = "center")
    print (f"Modified {ROW} cells from {ws.cell(row = 1, column = COL).value}!")


print("Changing the data!")

ws.cell(row = 1, column = 2).value = "Page Number"

for COL in range(1, ws.max_column+1):
    print("________________\n")
    print(f"Column Name: {ws.cell(row = 1, column = COL).value} \n")
    if COL == 5:
        print("This column is full of images!")
        image_loader = SheetImageLoader(ws)
        for ROW in range(2, ws.max_row+1):
            image = image_loader.get(f"E{ROW}")
            image.save(f"E{ROW}.png")  # Save image before workbook close
            print("Image Saved!")
    elif COL == 9:
        print("We are going to CHANGE the formula of this cell to count the length of the 7th (G) Row!")
        for ROW in range(2, ws.max_row+1):
            ws.cell(row = ROW, column = COL).value = f"=LEN(G{ROW})"
        print(f"\n{ROW} values were changed into a formula for G from {ws.cell(row = 1, column = COL).value}! \n")
    elif COL == 10:
        print("We are going to CHANGE the formula of this cell to count the length of the 8th (H) Row!")
        for ROW in range(2, ws.max_row+1):
            ws.cell(row = ROW, column = COL).value = f"=LEN(H{ROW})"
        print(f"\n{ROW} values were changed into a formula for H from {ws.cell(row = 1, column = COL).value}! \n")
    else:
        print("We are going to PURGE the values of this column")
        for ROW in range(2, ws.max_row+1):
            ws.cell(row = ROW, column = COL).value = None
        print(f"\n{ROW} values were DELETED from {ws.cell(row = 1, column = COL).value}! \n")

# Save workbook before closing
wb.save("ch155-001-018-9780323884051.xlsx")
print("File Saved!")

# Close workbook after all operations are done
wb.close()
print("File Closed!")
