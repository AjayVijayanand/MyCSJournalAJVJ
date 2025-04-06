import openpyxl as xl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl_image_loader import SheetImageLoader
import os
import logging

def extract_image(FILE, COLUMN, PATH):
    newpath = FILE[:-5]
    if not os.path.exists(newpath):
        os.makedirs(newpath)
    print("Extracting Images...")
    wb = xl.load_workbook(FILE)
    ws = wb.active
    for COL in COLUMN:
        image_loader = SheetImageLoader(ws)
        for ROW in range(2, ws.max_row+1):
            if image_loader.image_in(f"{COL}{ROW}"):
                image = image_loader.get(f"{COL}{ROW}")
                image.save(f"/{newpath}/image{ROW-1}.png")
                logging.info("Image Saved!")
    print("Images have been Extracted!")


def format_xl(FILE):
    wb = xl.load_workbook(FILE)
    ws = wb.active
    
    logging.info(f'The number of rows are {ws.max_row}')
    logging.info(f'The number of columns are {ws.max_column}')

    logging.info("Styling the data!")
    for COL in range(1, ws.max_column+1):
        for ROW in range(1, ws.max_row+1):
            ws.cell(row = ROW, column = COL).border = Border(
                top = Side(style = 'thin'),
                left = Side(style = 'thin'),
                right = Side(style = 'thin'),
                bottom = Side(style = 'thin')
            )
            if ROW != 1:
                ws.cell(row = ROW, column = COL).alignment = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
            else:
                ws.cell(row = ROW, column = COL).alignment = Alignment(horizontal = "center", vertical = "center")
        logging.info(f"Modified {ROW} cells from {ws.cell(row = 1, column = COL).value}!")

    logging.info("Changing the data!")
    ws.cell(row = 1, column = 2).value = "Page Number"
    for COL in range(1, ws.max_column+1):
        print("________________\n")
        print(f"Column Name: {ws.cell(row = 1, column = COL).value}")
        if ws.cell(row = 1, column = COL).value == "Thumbnail":
            print("This column is full of images! Use the extract_image function!")
        elif ws.cell(row = 1, column = COL).value == "Short Alt-text Length":
            print("We are going to CHANGE the formula of this cell to count the length of the Short Alt text!")
            for ROW in range(2, ws.max_row+1):
                ws.cell(row = ROW, column = COL).value = f"=LEN(G{ROW})"
            logging.info(f"{ROW} values were changed into a formula for G from {ws.cell(row = 1, column = COL).value}!")
        elif ws.cell(row = 1, column = COL).value == "Long Alt-text Length":
            print("We are going to CHANGE the formula of this cell to count the length of the Long Alt text!")
            for ROW in range(2, ws.max_row+1):
                ws.cell(row = ROW, column = COL).value = f"=LEN(H{ROW})"
            logging.info(f"{ROW} values were changed into a formula for H from {ws.cell(row = 1, column = COL).value}!")
        else:
            print("We are going to PURGE the values of this column")
            for ROW in range(2, ws.max_row+1):
                ws.cell(row = ROW, column = COL).value = None
            logging.info(f"{ROW} values were DELETED from {ws.cell(row = 1, column = COL).value}!")
    # Save workbook before closing
    wb.save(FILE)
    logging.info("File Saved!")

logging.basicConfig(filename="newfile.log",
                    format='%(asctime)s %(message)s',
                    filemode='w')

# Creating an object
logger = logging.getLogger()

# Setting the threshold of logger to DEBUG
# logger.setLevel(logging.DEBUG)

try:
    directory_in_str = "/Users/ajvj56/MyCSJournalAJVJ/VolumeX_FreeLancer/Python/ExcelParse"
    directory = os.fsencode(directory_in_str)
    for FILE in os.listdir(directory):
        filename = os.fsdecode(FILE)
        if filename.endswith(".xlsx"): 
            FILE = os.path.join(directory_in_str, filename)
            print(f"File currently: {filename} \n\n")
            format_xl(FILE)
            extract_image(FILE, "E", directory_in_str)
            continue
        else:
            continue
except Exception as err:
        print("An Error as occured, check the log!")
        logging.error(f"Unexpected {err=}, {type(err)=}")