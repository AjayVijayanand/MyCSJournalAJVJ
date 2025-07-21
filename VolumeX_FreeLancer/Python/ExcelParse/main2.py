import openpyxl as xl
from openpyxl.styles import Font, PatternFill
import os
import logging

def find_actual_max_row(ws):
    for row in reversed(range(1, ws.max_row + 1)):
        if any(cell.value not in (None, "") for cell in ws[row]):
            return row
    return 1  # At least the header row

def format_xl2(FILE, filename):
    wb = xl.load_workbook(FILE)
    ws = wb.active
    actual_max_row = find_actual_max_row(ws)

    print(f"File currently: {filename}")
    logging.info(f"File currently: {filename}")
    logging.info(f'The number of rows are {ws.max_row}')
    logging.info(f'The number of columns are {ws.max_column}')

    logging.info("Styling the data!")
    for COL in range(1, ws.max_column+1):
        for ROW in range(1, ws.max_row+1):
            if ROW == 1:
                ws.cell(row = ROW, column = COL).font = Font(
                    name='Calibri',
                    size=11,
                    color="000000",
                    bold="True"
                )
            else:
                ws.cell(row = ROW, column = COL).font = Font(
                    name='Calibri',
                    size=11,
                    color="000000"
                )

    for COL in range(1, ws.max_column+1):
        #Chapter Name/Number

        if ws.cell(row = 1, column = COL).value == "Chapter Name/Number":
            ExcLim = 0
            ExcLimRow = []
            for ROW in range(2, actual_max_row+1):
                if ws.cell(row = ROW, column = COL).value == None:
                    print(f"The ROW {ROW} does not have a Chapter Name/Number!")
                    ExcLim += 1
                    ExcLimRow.append(ROW)
                    ws.cell(row = ROW, column = 12).value = "NoCNN"
                    ws.cell(row = ROW, column = 12).fill = PatternFill(fill_type="solid",
                                                                       start_color="FF0000",
                                                                       end_color="FF0000") 
            if ExcLim > 0:
                print(f"\nThere are {ExcLim} rows that have no Chapter Name/Number in {filename}!")
                print(f"They are in the following Rows: {ExcLimRow}\n")

        #Page Number
        elif ws.cell(row = 1, column = COL).value == "Page Number":
            ExcLim = 0
            ExcLimRow = []
            for ROW in range(2, actual_max_row+1):
                if ws.cell(row = ROW, column = COL).value == None:
                    print(f"The ROW {ROW} does not have a Page Number!")
                    ExcLim += 1
                    ExcLimRow.append(ROW)
                    ws.cell(row = ROW, column = 13).value = "NoPN"
                    ws.cell(row = ROW, column = 13).fill = PatternFill(fill_type="solid",
                                                                       start_color="FF0000",
                                                                       end_color="FF0000") 
            if ExcLim > 0:
                print(f"\nThere are {ExcLim} rows that have no Page Number in {filename}!")
                print(f"They are in the following Rows: {ExcLimRow}\n")

        #Categories Formula Change
        elif ws.cell(row = 1, column = COL).value == "Category":
            for ROW in range(2, actual_max_row+1):
                ws.cell(row = ROW, column = COL).value = f"=IF(J{ROW}=0,\"Simple\",IF(J{ROW}<=700,\"Moderate\",\"Complex\"))"
            logging.info(f"{ROW} values were changed into a formula for H from {ws.cell(row = 1, column = COL).value}!")

        #Short Alt Text
        elif ws.cell(row = 1, column = COL).value == "Short Alt Text":
            ExcLim = 0
            ExcLimRow = []
            for ROW in range(2, actual_max_row+1):
                if ws.cell(row = ROW, column = COL).value == None:
                    print(f"The Short Alt Text has NONE on ROW {ROW}!")
                    ExcLim += 1
                    ExcLimRow.append(ROW)
                    ws.cell(row = ROW, column = 14).value = "STxtNONE"
                    ws.cell(row = ROW, column = 14).fill = PatternFill(fill_type="solid",
                                                                       start_color="FF0000",
                                                                       end_color="FF0000") 
                elif len(ws.cell(row = ROW, column = COL).value) > 200:
                    print(f"The Short Alt Text Exceeds 200 characters on ROW {ROW}!")
                    ExcLim += 1
                    ExcLimRow.append(ROW)
                    ws.cell(row = ROW, column = 14).value = "STxtExc"
                    ws.cell(row = ROW, column = 14).fill = PatternFill(fill_type="solid",
                                                                       start_color="FF0000",
                                                                       end_color="FF0000") 

            if ExcLim > 0:
                print(f"\nThere are {ExcLim} rows that have Short Alt Text that Exceeds 200 characters or MISSING short-alt texts in {filename}!")
                print(f"They are in the following Rows: {ExcLimRow}\n")

    # Save workbook before closing
    wb.save(FILE)
    logging.info("File Saved!")

logging.basicConfig(filename="newfile.log",
                    format='%(asctime)s %(message)s',
                    filemode='w')

# Creating an object
logger = logging.getLogger()

# Setting the threshold of logger to DEBUG
# logger.setLevel(logging.DEBUG)

try:
    directory_in_str = "/Users/ajvj56/MyCSJournalAJVJ/VolumeX_FreeLancer/Python/ExcelParse/9780128229712"
    directory = os.fsencode(directory_in_str)
    for FILE in os.listdir(directory):
        filename = os.fsdecode(FILE)
        if filename.endswith(".xlsx"): 
            FILE = os.path.join(directory_in_str, filename)
            format_xl2(FILE, filename)
            print("\n\n")
            continue
        else:
            continue
    print("All Files have been read!")
except Exception as err:
        print("An Error as occured, check the log!")
        logging.error(f"Unexpected {err=}, {type(err)=}")